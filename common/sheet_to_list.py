#!/usr/bin/env python3
import os
import sys
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import OrderedDict
from itertools import islice
import logging
import json
import getpass
logging.basicConfig(level=logging.INFO,
                    format="[%(asctime)s - %(levelname)s] %(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S")


def parse_arguments():

    parser = argparse.ArgumentParser(description="convert sample sheet to sample list with providing data dir")
    parser.add_argument("--sample-sheet", dest="sample_sheet", help="sample sheet file with xlsx format")
    parser.add_argument("--data-dir", dest="data_dir", help="directory all data lies")
    parser.add_argument("--out-dir", dest="out_dir", help="out dir")
    args = parser.parse_args()
    return args

def ws2df(ws):
    '''
    Transform worksheet to dataframe

    Args:
     ws:worksheet with sample massage

    Returns:
        dataframe of sample message
    '''
    data = ws.values
    cols = next(data)[0:]
    data = list(data)
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)
    return df

def find_fastq(sample_name, data_dir):
    '''
    Find the absolute path of fq file under the specified path based on the sample

    Args:
     sample_name: sample name (want to find her path)
     data_dir:absolute path of data

    Returns:
         absolute path of read1 and read2(fq)
    '''
    r1 = ''
    r2 = ''
    sample_name = sample_name + '_'
    abs = []
    fq_abs = []
    for root, list, files in os.walk(data_dir):
        for file in files:
            if sample_name in file:
                write = os.path.join(root, file)
                abs.append(write)
    # print(abs)
    for i in abs:
        if i.split(".")[-1] == 'gz':
            fq_abs.append(i)
    if len(fq_abs) == 2:
        if len(fq_abs[0]) == len(fq_abs[1]):
            for j in range(len(fq_abs[0])):
                if fq_abs[0][j] != fq_abs[1][j]:
                    if fq_abs[0][j] == "1":
                        r1 = fq_abs[0]
                        r2 = fq_abs[1]
                    else:
                        r1 = fq_abs[1]
                        r2 = fq_abs[0]
    return r1, r2

def convert_sample_sheet_to_list(xlsx, data_dir, out_dir):
    '''
    convert worksheet to list

    Args:
     xlsx: samplesheet
     data_dir:absolute path of data
     out_dir:absolute path for output list

    Returns:
    list of sample massage
    '''
    headers = ["sample_name", "fastq_r1", "fastq_r2"]
    wb = load_workbook(filename=xlsx)
    ws = wb[wb.sheetnames[0]]
    df = ws2df(ws)
    dna_sample_list = os.path.join(out_dir, "dna_sample_list")
    dna_sample_out = open(dna_sample_list, 'w')
    dna_sample_out.write("\t".join(headers) + "\n")
    rna_sample_list = os.path.join(out_dir, "rna_sample_list")
    rna_sample_out = open(rna_sample_list, 'w')
    rna_sample_out.write("\t".join(headers) + "\n")
    for index, row in df.iterrows():
        dna_sample_name = str(row["*样本编号(DNA)"]).strip()
        rna_sample_name = str(row["*样本编号(RNA)"]).strip()
        if dna_sample_name is None or dna_sample_name == "" or dna_sample_name == "None":
            continue
        if rna_sample_name is None or rna_sample_name == "" or rna_sample_name == "None":
            continue
        dna_r1, dna_r2 = find_fastq(dna_sample_name, data_dir)
        rna_r1, rna_r2 = find_fastq(rna_sample_name, data_dir)
        dna_sample_out.write("\t".join([dna_sample_name, dna_r1, dna_r2]) + "\n")
        rna_sample_out.write("\t".join([rna_sample_name, rna_r1, rna_r2]) + "\n")
    dna_sample_out.close()
    rna_sample_out.close()
    logging.info(f"dna samples' list {dna_sample_list} has been generated")
    logging.info(f"rna samples' list {rna_sample_list} has been generated")
        
if __name__ == '__main__':
    args = parse_arguments()
    data_dir = args.data_dir
    out_dir = args.out_dir
    convert_sample_sheet_to_list(args.sample_sheet, data_dir, out_dir)
